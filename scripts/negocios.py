"""
=== NEGOCIOS CARGO-GO ===

  1. Abre datos/negocios.xlsx en Excel
  2. Edita lo que quieras (nombre, link, estado, foto_local, etc.)
  3. Guarda el Excel
  4. Corre:  python scripts/negocios.py sync

  Eso es todo. Un comando sube TODO a Firestore.

  Columnas del Excel:
    - Estado: "activo" o "inactivo" (decide quien se ve en la app)
    - Link: URL del negocio (pagina, Maps, redes)
    - Foto Local: pon el nombre del archivo (ej: h01.jpg) y dejalo en datos/fotos/
                  o pon la ruta completa (ej: C:/Users/chule/fotos/tacos.jpg)
    - Foto URL: se llena solo al subir, no lo edites

  Otros comandos rapidos:
    python scripts/negocios.py ver            Ver todos
    python scripts/negocios.py ver tacos      Buscar
"""
import os, sys
import openpyxl
import firebase_admin
from firebase_admin import credentials, storage
from google.auth.transport.requests import AuthorizedSession
from google.oauth2 import service_account

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL = os.path.join(BASE, 'datos', 'negocios.xlsx')
FOTOS = os.path.join(BASE, 'datos', 'fotos')
SA = os.path.join(os.path.expanduser('~'), '.config', 'firebase', 'cargo-go-service-account.json')
PID = 'cargo-go-b5f77'
FS = f'https://firestore.googleapis.com/v1/projects/{PID}/databases/(default)/documents'

cred = credentials.Certificate(SA)
firebase_admin.initialize_app(cred, {'storageBucket': f'{PID}.firebasestorage.app'})
bkt = storage.bucket()
ses = AuthorizedSession(service_account.Credentials.from_service_account_file(
    SA, scopes=['https://www.googleapis.com/auth/datastore','https://www.googleapis.com/auth/cloud-platform']))

COLS = ['id','nombre','emoji','categoria','ciudad','zona','direccion','desc','horario','telefono','rating','pedidos','color_hex','link','foto_url','estado','plan','lat','lng','foto_local']

def _v(v):
    if isinstance(v,bool): return {'booleanValue':v}
    if isinstance(v,int): return {'integerValue':str(v)}
    if isinstance(v,float): return {'doubleValue':v}
    return {'stringValue':str(v) if v else ''}

def _leer():
    wb = openpyxl.load_workbook(EXCEL)
    data = []
    for nombre in ['Hidalgo', 'CDMX']:
        if nombre not in wb.sheetnames: continue
        ws = wb[nombre]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            n = {}
            for i, key in enumerate(COLS):
                val = row[i] if i < len(row) else ''
                n[key] = val if val is not None else ''
            if not n.get('plan'): n['plan'] = 'gratis'
            data.append(n)
    return data

def _guardar(data):
    wb = openpyxl.load_workbook(EXCEL)
    hid = [n for n in data if str(n.get('ciudad','')).lower() == 'tulancingo']
    cdmx = [n for n in data if str(n.get('ciudad','')).lower() == 'cdmx']
    for ws, rows in [(wb['Hidalgo'], hid), (wb['CDMX'], cdmx)]:
        for row_i, n in enumerate(rows, 2):
            for col_i, key in enumerate(COLS, 1):
                ws.cell(row=row_i, column=col_i, value=n.get(key, ''))
    wb.save(EXCEL)

def _subir_foto(nid, ruta):
    blob = bkt.blob(f'negocios/{nid}/foto.jpg')
    blob.upload_from_filename(ruta, content_type='image/jpeg')
    blob.make_public()
    return blob.public_url

def _resolver_foto(foto_local):
    """Busca la foto: primero ruta completa, luego en datos/fotos/"""
    if not foto_local: return None
    foto_local = str(foto_local).strip().strip('"').strip("'")
    if not foto_local: return None
    if os.path.isfile(foto_local): return foto_local
    en_fotos = os.path.join(FOTOS, foto_local)
    if os.path.isfile(en_fotos): return en_fotos
    return None

def ver(args):
    data = _leer()
    q = ' '.join(args).lower() if args else ''
    if q:
        data = [n for n in data if q in str(n.get('nombre','')).lower() or q in str(n.get('id','')).lower() or q in str(n.get('categoria','')).lower()]
    print(f'\n {"ID":<8} {"ON":<4} {"Nombre":<35} {"Ciudad":<13} {"Foto":<5} {"Link"}')
    print('-'*85)
    for n in data:
        est = 'SI' if n.get('estado')=='activo' else '--'
        foto = 'SI' if n.get('foto_url') else '--'
        lnk = str(n.get('link','')) or '-'
        if len(lnk)>25: lnk = lnk[:25]+'...'
        print(f' {str(n["id"]):<8} {est:<4} {str(n.get("nombre","?")):<35} {str(n.get("ciudad","?")):<13} {foto:<5} {lnk}')
    print(f'\n {len(data)} negocios')

def sync(args):
    data = _leer()
    print(f'\n=== Sincronizando {len(data)} negocios ===\n')
    ok = 0
    fotos_subidas = 0
    for n in data:
        nid = str(n['id'])
        # Subir foto si hay foto_local
        ruta = _resolver_foto(n.get('foto_local',''))
        if ruta:
            try:
                url = _subir_foto(nid, ruta)
                n['foto_url'] = url
                fotos_subidas += 1
                print(f'  📷 {nid}: foto subida')
            except Exception as e:
                print(f'  x {nid}: error foto - {e}')
        # Subir datos a Firestore (sin foto_local que es solo local)
        campos = {k:v for k,v in n.items() if k not in ('id','foto_local')}
        resp = ses.patch(
            f'{FS}/negocios/{nid}?{"&".join(f"updateMask.fieldPaths={k}" for k in campos)}',
            json={'fields': {k: _v(v) for k,v in campos.items()}})
        if resp.status_code == 200:
            ok += 1
            est = 'ON' if n.get('estado')=='activo' else 'OFF'
            print(f'  + {nid}: {n.get("nombre","?")} [{est}]')
        else:
            print(f'  x {nid}: error {resp.status_code}')
    # Guardar Excel con foto_urls actualizados
    _guardar(data)
    print(f'\n{ok}/{len(data)} negocios actualizados')
    if fotos_subidas: print(f'{fotos_subidas} fotos subidas')
    print('Listo!')

CMDS = {'ver': ver, 'sync': sync}

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit()
    c = sys.argv[1].lower()
    if c in CMDS:
        CMDS[c](sys.argv[2:])
    else:
        print(f'"{c}" no existe\n')
        print(__doc__)
